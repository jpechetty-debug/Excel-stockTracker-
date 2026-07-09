import openpyxl

file_path = '../IOS_Master_Tracker_Filled_13.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['Master Universe']

headers = {cell.value: i for i, cell in enumerate(ws[1])}
ticker_col = headers['Ticker']
company_col = headers['Company']

existing_tickers = set()
for row in ws.iter_rows(min_row=2):
    t = row[ticker_col].value
    if t:
        existing_tickers.add(t)

new_data = [
    ("Sterlite Technologies", "STLTECH"),
    ("MTAR Technologies", "MTARTECH"),
    ("Garware Hi-Tech Films", "GRWRHITECH"),
    ("Cupid", "CUPID"),
    ("Rubicon Research", "RUBICON"),
    ("Sun Pharma Adv. Research", "SPARC"),
    ("Thangamayil Jewellery", "THANGAMAYL"),
    ("Park Medi World", "PARKHOSPS"),
    ("Paras Defence", "PARAS"),
    ("Sansera Engineering", "SANSERA"),
    ("Atlanta Electricals", "ATLANTAELE"),
    ("Shilpa Medicare", "SHILPAMED"),
    ("Avalon Technologies", "AVALON"),
    ("Balaji Amines", "BALAMINES"),
    ("Astra Microwave Products", "ASTRAMICRO")
]

# Find the first empty row based on the Ticker column
start_row = 2
while ws.cell(row=start_row, column=ticker_col+1).value is not None:
    start_row += 1

added = 0
for company, symbol in new_data:
    ticker = f"{symbol}.NS"
    if ticker not in existing_tickers:
        ws.cell(row=start_row, column=ticker_col+1, value=ticker)
        ws.cell(row=start_row, column=company_col+1, value=company)
        start_row += 1
        added += 1

wb.save(file_path)
print(f"Added {added} new tickers.")
