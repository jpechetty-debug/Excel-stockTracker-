import openpyxl

wb = openpyxl.load_workbook('../IOS_Master_Tracker_Filled_13.xlsx', data_only=True)
ws = wb['Master Universe']

headers = {}
header_row = 1
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10)):
    vals = [c.value for c in row if isinstance(c.value, str)]
    if 'Ticker' in vals or 'Symbol' in vals:
        headers = {c.value: idx for idx, c in enumerate(row)}
        header_row = i + 1
        break

for ticker in ['SBIN.NS', 'PFC.NS']:
    ticker_col = headers.get('Ticker') or headers.get('Symbol')
    row = next((r for r in ws.iter_rows(min_row=header_row+1) if r[ticker_col].value == ticker), None)
    if row:
        inv_score = row[headers.get('Investment Score', 0)].value if 'Investment Score' in headers else 'N/A'
        risk_score = row[headers.get('Risk Score', 0)].value if 'Risk Score' in headers else 'N/A'
        bus_score = row[headers.get('Business Score', 0)].value if 'Business Score' in headers else 'N/A'
        fin_risk = row[headers.get('Financial Risk', 0)].value if 'Financial Risk' in headers else 'N/A'
        tot_risk = row[headers.get('Total Risk', 0)].value if 'Total Risk' in headers else 'N/A'
        gate = row[headers.get('Evidence Gate', 0)].value if 'Evidence Gate' in headers else 'N/A'
        sector = row[headers.get('Sector', 0)].value if 'Sector' in headers else 'N/A'
        industry = row[headers.get('Industry', 0)].value if 'Industry' in headers else 'N/A'
        gnpa = row[headers.get('Gross NPA %', 0)].value if 'Gross NPA %' in headers else 'N/A'
        car = row[headers.get('CAR %', 0)].value if 'CAR %' in headers else 'N/A'
        print(f'{ticker}: Gate={gate}, Sector={sector}, Industry={industry}, GNPA={gnpa}, CAR={car}, Inv={inv_score}, Risk={risk_score}, Bus={bus_score}')
        
print("Headers:", list(headers.keys()))
