import pytest
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from infrastructure.excel.reader import ExcelReader
from infrastructure.excel.writer import ExcelWriter
from domain.models import CompanyData, Metric, Provenance
from datetime import datetime
from unittest.mock import MagicMock

@pytest.fixture
def mock_config():
    config = MagicMock()
    config.column_map = {
        "sheets": {"master": "Master Universe"},
        "keys": {"primary": "Ticker", "company": "Company"},
        "zones": {
            "system": [
                {"field": "Current Price", "column": "Current Price"},
                {"field": "Sector", "column": "Sector"}
            ],
            "manual": [
                {"field": "Manual Thesis", "column": "Manual Thesis"}
            ]
        }
    }
    return config

@pytest.fixture
def complex_workbook(tmp_path):
    file_path = tmp_path / "Master_Tracker_Test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Universe"
    
    headers = ["Ticker", "Company", "Sector", "Current Price", "Manual Thesis", "Formula Field"]
    ws.append(headers)
    
    # Row 2 (Data)
    ws.append(["TEST", "Test Co", "Tech", 100.0, "Strong Buy", ""])
    ws["F2"] = "=D2*2"
    
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    bold_font = Font(bold=True)
    ws["E2"].fill = yellow_fill
    ws["D2"].font = bold_font
    
    wb.save(file_path)
    return file_path

def test_workbook_integrity_preservation(mock_config, complex_workbook, tmp_path):
    reader = ExcelReader(config=mock_config)
    companies, raw_data = reader.read_companies(complex_workbook)
    
    assert len(companies) == 1
    c = companies[0]
    assert c.ticker == "TEST"
    
    prov = Provenance("mock", "mock", datetime.now(), 100, "1", "1")
    c.metrics["Current Price"] = Metric(150.0, prov)
    c.metrics["Sector"] = Metric("Finance", prov)
    from infrastructure.excel.planner import UpdatePlanner
    planner = UpdatePlanner(config=mock_config)
    plan = planner.create_plan(companies, raw_data)
    
    writer = ExcelWriter(config=mock_config)
    out_path = tmp_path / "Master_Tracker_Out.xlsx"
    writer.update_workbook(complex_workbook, plan, out_path, tmp_path)
    
    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    ws = wb["Master Universe"]
    
    assert ws["D2"].value == 150.0
    assert ws["C2"].value == "Finance"
    assert ws["D2"].font.bold is True
    assert ws["E2"].value == "Strong Buy"
    assert ws["E2"].fill.start_color.index == '00FFFF00'
    assert ws["F2"].value == "=D2*2"
