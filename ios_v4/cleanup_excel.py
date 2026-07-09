import openpyxl
import os
import shutil

INPUT_FILE = "../IOS_Master_Tracker_Filled_13.xlsx"
OUTPUT_FILE = "../IOS_Master_Tracker_Filled_13_Clean.xlsx"

def main():
    print(f"Loading {INPUT_FILE}...")
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=False)
    
    master_ws = wb["Master Universe"]
    fin_ws = wb["Financial Data"] if "Financial Data" in wb.sheetnames else None

    # Helper: get column index by header name
    def get_col_idx(ws, header_name):
        for col_idx in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col_idx).value == header_name:
                return col_idx
        return None

    # 1 & 2: Reconcile ROE and delete from Financial Data
    if fin_ws:
        fin_company_col = get_col_idx(fin_ws, "Company")
        fin_roe_col = get_col_idx(fin_ws, "ROE %")
        
        master_company_col = get_col_idx(master_ws, "Company")
        master_roe_col = get_col_idx(master_ws, "ROE %")
        
        if fin_company_col and fin_roe_col and master_company_col and master_roe_col:
            # Build Company -> ROE map from Financial Data
            roe_map = {}
            for row in range(2, fin_ws.max_row + 1):
                company = fin_ws.cell(row=row, column=fin_company_col).value
                roe = fin_ws.cell(row=row, column=fin_roe_col).value
                if company:
                    roe_map[company] = roe
            
            # Apply to Master Universe
            updates = 0
            for row in range(2, master_ws.max_row + 1):
                company = master_ws.cell(row=row, column=master_company_col).value
                if company in roe_map:
                    master_ws.cell(row=row, column=master_roe_col).value = roe_map[company]
                    updates += 1
            print(f"Reconciled ROE for {updates} companies.")
            
            # Delete ROE column from Financial Data
            fin_ws.delete_cols(fin_roe_col)
            print("Deleted 'ROE %' column from 'Financial Data' sheet.")

    # 3. Clear stray formulas
    cols_to_clear = [
        "Target Allocation %", 
        "Margin of Safety %", 
        "EPS Implied (Rs, Price/PE)", 
        "Justified P/E"
    ]
    
    cleared = 0
    for col_name in cols_to_clear:
        col_idx = get_col_idx(master_ws, col_name)
        if col_idx:
            for row in range(2, master_ws.max_row + 1):
                cell = master_ws.cell(row=row, column=col_idx)
                if isinstance(cell.value, str) and str(cell.value).startswith("="):
                    cell.value = None
                    cleared += 1
    print(f"Cleared {cleared} stray formulas.")

    # 4. Add missing columns
    cols_to_add = [
        "Gross NPA %", "Net NPA %", "CAR %", "PCR %", 
        "Business Risk", "Financial Risk", "Valuation Risk", "Governance Risk", "Total Risk"
    ]
    
    next_col = master_ws.max_column + 1
    added = 0
    for col_name in cols_to_add:
        if not get_col_idx(master_ws, col_name):
            master_ws.cell(row=1, column=next_col).value = col_name
            next_col += 1
            added += 1
    print(f"Added {added} missing columns.")

    print(f"Saving to {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)
    print("Done!")

if __name__ == "__main__":
    main()
