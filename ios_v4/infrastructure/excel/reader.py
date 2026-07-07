"""
Excel Reader Module
Reads the Master Universe and extracts CompanyData objects.
"""

import openpyxl
from pathlib import Path
from typing import List, Dict, Optional, Any
from domain.models import CompanyData
from config.config_loader import ConfigLoader
from infrastructure.logging.logger import logger


class ExcelReader:
    """Reads companies from the Excel workbook."""

    def __init__(self, config: ConfigLoader):
        self.config = config
        self.master_sheet_name = config.column_map.get("sheets", {}).get("master", "Master Universe")
        self.primary_key = config.column_map.get("keys", {}).get("primary", "Ticker")
        self.company_key = config.column_map.get("keys", {}).get("company", "Company")

    def read_companies(self, filepath: Path | str, limit: Optional[int] = None, tickers: Optional[List[str]] = None) -> tuple[List[CompanyData], Dict[str, Dict[str, Any]]]:
        """Reads the Excel file and extracts a list of CompanyData objects and existing system data."""
        path = Path(filepath)
        if not path.exists():
            raise FileNotFoundError(f"Workbook not found: {path}")

        logger.info(f"Reading workbook: {path}")
        wb = openpyxl.load_workbook(path, data_only=True)
        
        if self.master_sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{self.master_sheet_name}' not found in {path}")

        ws = wb[self.master_sheet_name]
        return self._parse_sheet(ws, limit, tickers)

    def _parse_sheet(self, ws, limit: Optional[int] = None, tickers: Optional[List[str]] = None) -> tuple[List[CompanyData], Dict[str, Dict[str, Any]]]:
        """Parses the worksheet rows into CompanyData objects and existing system data."""
        companies = []
        existing_data = {}
        headers = {cell.value: idx for idx, cell in enumerate(ws[1]) if cell.value}
        
        if self.primary_key not in headers:
            raise ValueError(f"Primary key '{self.primary_key}' not found in headers.")

        ticker_col = headers[self.primary_key]
        company_col = headers.get(self.company_key)

        system_zones = self.config.column_map.get("zones", {}).get("system", [])
        field_col_map = {item["field"]: item["column"] for item in system_zones if "field" in item and "column" in item}

        for row in ws.iter_rows(min_row=2, values_only=True):
            ticker = row[ticker_col]
            if not ticker:
                continue  # Skip empty rows
                
            ticker_str = str(ticker).strip()
            
            # Apply tickers filter
            if tickers and ticker_str not in tickers:
                continue

            company_name = row[company_col] if company_col is not None else ticker_str
            companies.append(CompanyData(ticker=ticker_str, company_name=str(company_name)))
            
            # Extract existing data for the planner
            existing = {}
            for field, col_name in field_col_map.items():
                if col_name in headers:
                    val = row[headers[col_name]]
                    existing[field] = val
            existing_data[ticker_str] = existing
            
            # Apply limit filter
            if limit and len(companies) >= limit:
                break
            
        logger.info(f"Extracted {len(companies)} companies from {ws.title}.")
        return companies, existing_data
