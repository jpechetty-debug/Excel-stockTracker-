"""
Excel Writer Module
Updates the System Zone while strictly protecting the Manual Zone.
"""

import openpyxl
from pathlib import Path
from typing import List, Set
from domain.models import CompanyData
from config.config_loader import ConfigLoader
from infrastructure.logging.logger import logger, log_audit_trail


class ExcelWriter:
    """Updates the workbook with strict boundary enforcement."""

    def __init__(self, config: ConfigLoader):
        self.config = config
        self.master_sheet_name = config.column_map.get("sheets", {}).get("master", "Master Universe")
        self.primary_key = config.column_map.get("keys", {}).get("primary", "Ticker")
        
        # Build strict allowed columns set
        self.allowed_columns: Set[str] = set()
        system_zones = config.column_map.get("zones", {}).get("system", [])
        for item in system_zones:
            if "column" in item:
                self.allowed_columns.add(item["column"])

    def update_workbook(self, filepath: Path | str, plan: 'WorkbookUpdatePlan', output_path: Path | str, reports_dir: Path | str) -> None:
        """Updates allowed cells, writes diff report, and saves to output_path."""
        logger.info(f"Opening workbook for writing: {filepath}")
        wb = openpyxl.load_workbook(filepath)
        ws = wb[self.master_sheet_name]

        headers = {cell.value: idx for idx, cell in enumerate(ws[1]) if cell.value}
        ticker_col = headers.get(self.primary_key)
        
        if ticker_col is None:
            raise ValueError(f"Primary key '{self.primary_key}' not found in headers.")

        # Map field -> column_name
        field_to_col = {}
        system_zones = self.config.column_map.get("zones", {}).get("system", [])
        for item in system_zones:
            if "field" in item and "column" in item:
                field_to_col[item["field"]] = item["column"]

        # Map Ticker -> Row Index
        row_map = {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            ticker_cell = row[ticker_col]
            if ticker_cell.value:
                row_map[str(ticker_cell.value)] = row_idx

        import csv
        from infrastructure.excel.planner import UpdateStatus
        
        Path(reports_dir).mkdir(parents=True, exist_ok=True)
        diff_path = Path(reports_dir) / "workbook_diff.csv"
        
        with open(diff_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Sheet", "Ticker", "Field", "Column", "Old Value", "New Value", "Status", "Reason"])
            
            for action in plan.actions:
                col_name = field_to_col.get(action.field)
                if not col_name or col_name not in headers:
                    if action.status == UpdateStatus.UPDATED:
                        action.status = UpdateStatus.FAILED
                        action.reason = f"Column '{col_name}' not found in sheet"
                
                # Write to Excel if UPDATED
                if action.status == UpdateStatus.UPDATED:
                    row_idx = row_map.get(action.ticker)
                    if row_idx:
                        col_idx = headers[col_name] + 1
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = action.new_value
                        
                        log_audit_trail(
                            ticker=action.ticker,
                            field=action.field,
                            old_val=action.old_value,
                            new_val=action.new_value,
                            source=action.provider,
                            confidence=action.confidence
                        )
                    else:
                        action.status = UpdateStatus.FAILED
                        action.reason = "Ticker row not found in workbook"
                
                # Write to Diff CSV
                writer.writerow([
                    self.master_sheet_name,
                    action.ticker,
                    action.field,
                    col_name or "N/A",
                    action.old_value,
                    action.new_value,
                    action.status.name,
                    action.reason
                ])

        logger.info(f"Saving updated workbook to: {output_path}")
        wb.save(output_path)
        
        # Validation Step
        self._validate_workbook(output_path)
        
    def _validate_workbook(self, filepath: Path | str) -> None:
        """Post-save integrity check to ensure workbook opens successfully."""
        try:
            wb = openpyxl.load_workbook(filepath, data_only=False)
            if self.master_sheet_name not in wb.sheetnames:
                logger.error("Validation Failed: Master sheet missing after save!")
                raise ValueError("Corrupt save: Master sheet missing")
            logger.info("Workbook post-save validation passed.")
        except Exception as e:
            logger.error(f"Workbook corrupted during save: {e}")
            raise
