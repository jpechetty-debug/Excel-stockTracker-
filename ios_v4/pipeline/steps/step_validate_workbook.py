"""
Step: Validate Workbook
Post-write sanity pass over the saved Excel file. Runs AFTER StepWriteExcel so it
checks what actually landed in the workbook (including pre-existing manual
columns the pipeline itself never touches, like "Data Confidence (%)") rather
than only the in-memory values the pipeline computed this run.

This is intentionally non-fatal: it logs clear WARNING/ERROR entries so issues
are visible in the run log and reports, but doesn't halt the pipeline, since a
data-quality flag on one ticker shouldn't block writing results for the other
35.
"""
from pathlib import Path
from openpyxl import load_workbook
from pipeline.context import ExecutionContext, Severity
from pipeline.runner import PipelineStep


# Columns that must be numeric within [0, 100] whenever populated.
CONFIDENCE_COLUMNS = [
    "Data Confidence (%)",
    "Business Confidence (%)",
    "Investment Confidence (%)",
    "Risk Confidence (%)",
]


class StepValidateWorkbook:

    @property
    def name(self) -> str:
        return "Validate Workbook (Post-Write)"

    def validate(self, context: ExecutionContext) -> bool:
        if not context.artifacts.final_output_path:
            context.log("No output path recorded; skipping workbook validation.", Severity.WARNING)
            return False
        return True

    def execute(self, context: ExecutionContext) -> bool:
        try:
            path = context.artifacts.final_output_path
            if not Path(path).exists():
                context.log(f"Cannot validate - output file not found: {path}", Severity.WARNING)
                return True  # non-fatal

            wb = load_workbook(path, data_only=True)
            if "Master Universe" not in wb.sheetnames:
                context.log("No 'Master Universe' sheet found; skipping validation.", Severity.WARNING)
                return True

            ws = wb["Master Universe"]
            headers = [c.value for c in ws[1]]
            col = {h: i + 1 for i, h in enumerate(headers) if h is not None}

            portfolio_config = context.config.get("portfolio", {})
            max_position_pct = portfolio_config.get("max_position_size", 0.10) * 100

            issues = []

            def cell(row, colname):
                if colname not in col:
                    return None
                return ws.cell(row=row, column=col[colname]).value

            for r in range(2, ws.max_row + 1):
                company = ws.cell(row=r, column=2).value
                if not company:
                    continue

                # Rule 1: confidence columns must be numeric 0-100, or blank.
                for cname in CONFIDENCE_COLUMNS:
                    if cname not in col:
                        continue
                    v = cell(r, cname)
                    if v is None or v == "":
                        continue
                    if not isinstance(v, (int, float)):
                        issues.append(f"[{company}] {cname} = {v!r} is not numeric.")
                    elif not (0 <= v <= 100):
                        issues.append(f"[{company}] {cname} = {v} is outside 0-100.")

                # Rule 2: Target Allocation must not exceed the configured max
                # position size. Stored as strings like "10%", "7.5%",
                # "Watchlist only", "Pending Score", "Not Eligible (Failed Gate)".
                alloc_raw = cell(r, "Target Allocation %")
                if isinstance(alloc_raw, str) and alloc_raw.endswith("%"):
                    try:
                        alloc_pct = float(alloc_raw.rstrip("%"))
                        if alloc_pct > max_position_pct + 1e-9:
                            issues.append(
                                f"[{company}] Target Allocation {alloc_pct}% exceeds "
                                f"configured max position size {max_position_pct}%."
                            )
                    except ValueError:
                        pass

                # Rule 3 & 4: Intrinsic Value / Buy Price must not be negative,
                # and Buy Price must not exceed Intrinsic Value when both exist
                # (Buy Price = Intrinsic Value x (1 - required margin of safety),
                # so it should always sit at or below Intrinsic Value).
                iv = cell(r, "Intrinsic Value")
                bp = cell(r, "Buy Price")
                if isinstance(iv, (int, float)) and iv < 0:
                    issues.append(f"[{company}] Intrinsic Value is negative: {iv}")
                if isinstance(bp, (int, float)) and bp < 0:
                    issues.append(f"[{company}] Buy Price is negative: {bp}")
                if isinstance(iv, (int, float)) and isinstance(bp, (int, float)):
                    if bp > iv + 1e-6:
                        issues.append(
                            f"[{company}] Buy Price ({bp}) exceeds Intrinsic Value ({iv}) - "
                            f"should never happen given Buy Price = IV x (1 - MoS)."
                        )

            if issues:
                context.log(f"Workbook validation found {len(issues)} issue(s):", Severity.WARNING)
                for issue in issues:
                    context.log(f"  - {issue}", Severity.WARNING)
            else:
                context.log("Workbook validation passed: no issues found.", Severity.INFO)

            return True

        except Exception as e:
            context.log(f"Workbook validation step failed to run: {str(e)}", Severity.WARNING)
            return True  # non-fatal - don't let the validator itself break the pipeline

    def rollback(self, context: ExecutionContext) -> None:
        pass
