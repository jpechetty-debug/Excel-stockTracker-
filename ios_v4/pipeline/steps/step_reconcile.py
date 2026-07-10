"""
Step: Reconcile Financial Data
Reads the 'Financial Data' sheet (manually-researched, cited figures) and cross-checks
its ROE against what the automated engine fetched, surfacing conflicts as a durable
Excel column instead of a one-time manual patch that the next pipeline run would
silently undo.

Note: 'Financial Data' is otherwise never read by the pipeline - it's a pure
human-facing reference sheet today, so this is the first place it connects to
the rest of the engine.
"""
import re
from openpyxl import load_workbook
from pipeline.context import ExecutionContext, Severity
from pipeline.runner import PipelineStep
from domain.engines.reconciliation import ReconciliationEngine


def _normalize_name(name: str) -> str:
    if not name:
        return ""
    # Strip parenthetical tickers/abbreviations e.g. "(CAMS)" and lowercase,
    # so "State Bank of India" and "State Bank Of India Ltd" style variants match.
    name = re.sub(r"\(.*?\)", "", name)
    name = re.sub(r"\bltd\.?\b|\blimited\b", "", name, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", name).strip().lower()


class StepReconcile:

    @property
    def name(self) -> str:
        return "Reconcile Financial Data Sheet"

    def validate(self, context: ExecutionContext) -> bool:
        if not context.artifacts.financials:
            context.log("Missing financials - skipping reconciliation.", Severity.WARNING)
            return False
        return True

    def execute(self, context: ExecutionContext) -> bool:
        try:
            master_file = context.input_file or context.config.get("excel", {}).get("master_file")
            if not master_file:
                context.log("No master_file configured - skipping reconciliation.", Severity.WARNING)
                return True  # non-fatal, this is a supplementary check

            wb = load_workbook(master_file, data_only=True)
            if "Financial Data" not in wb.sheetnames:
                context.log("No 'Financial Data' sheet found - skipping reconciliation.", Severity.INFO)
                return True

            ws = wb["Financial Data"]
            headers = {cell.value: idx for idx, cell in enumerate(ws[1], 1) if cell.value}
            if "Company" not in headers or "ROE %" not in headers:
                context.log("'Financial Data' sheet missing Company/ROE % columns - skipping.", Severity.WARNING)
                return True

            company_col = headers["Company"]
            roe_col = headers["ROE %"]

            # Manually-researched ROE, keyed by normalized company name.
            # Last non-empty entry wins if a company appears in multiple rows/quarters.
            reference_roe = {}
            for row in ws.iter_rows(min_row=2, values_only=False):
                name = row[company_col - 1].value
                roe = row[roe_col - 1].value
                if name and isinstance(roe, (int, float)):
                    reference_roe[_normalize_name(str(name))] = float(roe)

            engine = ReconciliationEngine(variance_threshold=0.15)
            raw_companies = context.artifacts.raw_companies
            financials = context.artifacts.financials
            checked = 0

            for company in raw_companies:
                ticker = company.ticker
                if ticker not in financials:
                    continue
                key = _normalize_name(company.company_name)
                if key not in reference_roe:
                    continue

                engine_roe_frac = financials[ticker].breakdown.get("roe")
                if engine_roe_frac is None:
                    continue
                engine_roe_pct = float(engine_roe_frac) * 100
                reference_roe_pct = reference_roe[key]

                result = engine.check_metric("ROE", engine_roe_pct, reference_roe_pct)
                context.artifacts.reconciliation[ticker] = result
                checked += 1

                if result.warnings:
                    context.log(f"[{ticker}] {result.reasons[0]}", Severity.WARNING)

            context.log(f"Cross-checked ROE for {checked} companies against Financial Data sheet.", Severity.INFO)
            return True

        except Exception as e:
            # Non-fatal by design - this is a supplementary data-quality check,
            # not core to producing a valuation. A failure here shouldn't block the run.
            context.log(f"Reconciliation check failed (non-fatal): {str(e)}", Severity.WARNING)
            return True

    def rollback(self, context: ExecutionContext) -> None:
        context.artifacts.reconciliation = {}
